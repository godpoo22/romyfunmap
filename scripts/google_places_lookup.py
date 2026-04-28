"""
google_places_lookup.py
=======================
從景點/餐廳名稱清單批次查詢 Google Places API，輸出對照表 Excel。

使用方式
--------
1. 把你的 API Key 填入下方 API_KEY
2. 在 INPUT_FILE 指定輸入檔（TSV 或純文字清單，見說明）
3. 執行：python google_places_lookup.py
4. 查看輸出的 output_lookup.xlsx

輸入格式（兩種）
----------------
A) 直接用現有 TSV（會自動讀 title + district 欄）
   INPUT_FORMAT = "tsv"
   INPUT_FILE   = "data/taipei_attractions.tsv"

B) 自己建一個純文字清單（每行一筆，名稱,縣市）
   INPUT_FORMAT = "list"
   INPUT_FILE   = "my_list.txt"
   範例內容：
     國立海洋生物博物館,屏東縣
     科學博物館,台中市
     奇美博物館,台南市
"""

import csv
import os
import sys
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── 設定區 ──────────────────────────────────────────
API_KEY      = "YOUR_API_KEY_HERE"   # ← 填入你的 Google API Key
INPUT_FORMAT = "tsv"                 # "tsv" 或 "list"
INPUT_FILE   = "data/taipei_attractions.tsv"
OUTPUT_FILE  = "output_lookup.xlsx"
SEARCH_REGION = "TW"                 # 搜尋地區（台灣）
LANGUAGE      = "zh-TW"
# ────────────────────────────────────────────────────


def text_search(name, city):
    """用 Places Text Search 找到最相關的店家"""
    query = f"{name} {city}"
    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {
        "query": query,
        "region": SEARCH_REGION,
        "language": LANGUAGE,
        "key": API_KEY,
    }
    resp = requests.get(url, params=params, timeout=10)
    data = resp.json()
    if data.get("status") == "OK" and data.get("results"):
        return data["results"][0]  # 取第一筆最相關結果
    return None


def place_details(place_id):
    """用 Place ID 取得詳細資料（電話、地址、座標等）"""
    url = "https://maps.googleapis.com/maps/api/place/details/json"
    params = {
        "place_id": place_id,
        "fields": "name,formatted_address,formatted_phone_number,geometry,types,url",
        "language": LANGUAGE,
        "key": API_KEY,
    }
    resp = requests.get(url, params=params, timeout=10)
    data = resp.json()
    if data.get("status") == "OK":
        return data.get("result", {})
    return {}


def load_input():
    """讀取輸入，回傳 [(原始名稱, 縣市/區, 原始地址, 原始lat, 原始lng), ...]"""
    items = []
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(script_dir, INPUT_FILE)

    if INPUT_FORMAT == "tsv":
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                title   = row.get("title", "").strip()
                district = row.get("district", "").strip()
                address = row.get("address", "").strip()
                # TSV 沒有獨立 lat/lng 欄位，略過
                items.append((title, district, address, "", ""))
    elif INPUT_FORMAT == "list":
        with open(filepath, encoding="utf-8-sig") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split(",", 1)
                name = parts[0].strip()
                city = parts[1].strip() if len(parts) > 1 else ""
                items.append((name, city, "", "", ""))
    else:
        print("ERROR: INPUT_FORMAT 必須是 'tsv' 或 'list'")
        sys.exit(1)

    return items


def diff_flag(a, b):
    """簡單比對兩個字串，有差異回傳 '⚠ 差異'"""
    a2 = a.strip().replace(" ", "").replace("　", "")
    b2 = b.strip().replace(" ", "").replace("　", "")
    return "⚠ 差異" if a2 and b2 and a2 not in b2 and b2 not in a2 else ""


def main():
    if API_KEY == "YOUR_API_KEY_HERE":
        print("請先在腳本頂端填入你的 Google API Key！")
        sys.exit(1)

    items = load_input()
    print(f"共 {len(items)} 筆，開始查詢 Google Places API...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Places 對照"

    # 標題列
    headers = [
        "原始名稱", "輸入縣市/區",
        "Google 名稱", "Google 地址", "Google 電話",
        "Google Lat", "Google Lng",
        "原始地址", "名稱差異", "地址差異",
        "Google 類型", "Google Maps 連結", "Place ID",
    ]
    header_fill = PatternFill("solid", fgColor="1D6A54")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # 警告用底色
    warn_fill  = PatternFill("solid", fgColor="FFF3CD")
    error_fill = PatternFill("solid", fgColor="FFDCDC")
    ok_fill    = PatternFill("solid", fgColor="D4EDDA")

    for row_idx, (orig_name, city, orig_addr, orig_lat, orig_lng) in enumerate(items, 2):
        print(f"  [{row_idx-1}/{len(items)}] {orig_name}...")
        g_name = g_addr = g_phone = g_lat = g_lng = g_types = g_url = place_id = ""

        try:
            result = text_search(orig_name, city)
            if result:
                place_id = result.get("place_id", "")
                g_name   = result.get("name", "")
                g_addr   = result.get("formatted_address", "")
                loc      = result.get("geometry", {}).get("location", {})
                g_lat    = str(loc.get("lat", ""))
                g_lng    = str(loc.get("lng", ""))
                g_types  = ", ".join(result.get("types", [])[:5])

                # 取詳細資料（電話 + Maps 連結）
                if place_id:
                    details  = place_details(place_id)
                    g_phone  = details.get("formatted_phone_number", "")
                    g_url    = details.get("url", "")
                    if not g_name:
                        g_name = details.get("name", "")
                    if not g_addr:
                        g_addr = details.get("formatted_address", "")

            time.sleep(0.2)  # 避免超過 API 速率限制

        except Exception as e:
            g_name = f"ERROR: {e}"

        name_diff = diff_flag(orig_name, g_name)
        addr_diff = diff_flag(orig_addr, g_addr)

        row_data = [
            orig_name, city,
            g_name, g_addr, g_phone,
            g_lat, g_lng,
            orig_addr, name_diff, addr_diff,
            g_types, g_url, place_id,
        ]

        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

        # 上色
        if not g_name:
            fill = error_fill   # 找不到
        elif name_diff or addr_diff:
            fill = warn_fill    # 有差異
        else:
            fill = ok_fill      # 一致

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    # 欄寬
    col_widths = [20, 12, 24, 40, 16, 10, 10, 36, 10, 10, 30, 40, 30]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.freeze_panes = "A2"

    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(script_dir, OUTPUT_FILE)
    wb.save(out_path)
    print(f"\n完成！輸出到：{out_path}")
    print("顏色說明：🟢 一致  🟡 有差異  🔴 找不到")


if __name__ == "__main__":
    main()
